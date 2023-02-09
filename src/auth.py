import openpyxl
import json


def valid_usn(usn, file_path="../data/usn.json"):
    data = json.load(open(file_path))
    if int(usn) in data["data"]:
        return True
    else:
        return False
    """wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    print(sheet.max_row)

    for i in range(1, sheet.max_row + 1):
        cell = sheet.cell(row=i, column=1)
        if cell.value == usn:
            return True
    return False"""


def done(usn, file_path="../data/done.json"):
    data = json.load(open(file_path))
    if int(usn) in data["data"]:
        return True
    else:
        return False


def addDone(usn, file_path="../data/done.json"):
    data = json.load(open(file_path))
    data["data"].append(int(usn))
    with open(file_path, "w") as f:
        json.dump(data, f, indent=4)


def add_to_excel(usn, file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    a = sheet.max_row + 1
    cell = sheet.cell(row=a, column=1)
    cell.value = usn
    wb.save(file_path)
